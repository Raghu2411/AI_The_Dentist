import json
import os
import time
import re 
from groq import Groq, GroqError
from collections import deque
from tqdm import tqdm
import config_and_prompts
import generate_grounding_caption

from PIL import Image as PILImage
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.pagesizes import portrait, A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import black
from openpyxl import Workbook, load_workbook


# Configuration
INPUT_JSON_FILE = 'C:/Users/USER/Desktop/OPG-segmentation/Part 2/final/Dental_Dissertation/Report Generation/ensemble_output/_ensemble_annotations.json'
OUTPUT_DIR = 'C:/Users/USER/Desktop/OPG-segmentation/Part 2/final/Dental_Dissertation/Report Generation/medical_reports'
MODEL_ID = 'qwen/qwen3-32b'

SOURCE_IMAGE_DIR = 'C:/Users/USER/Desktop/OPG-segmentation/Part 2/final/Dental_Dissertation/Report Generation/ensemble_output'
PDF_OUTPUT_DIR = 'C:/Users/USER/Desktop/OPG-segmentation/Part 2/final/Dental_Dissertation/Report Generation/medical_PDFs'

EVALUATION_EXCEL_FILE = 'C:/Users/USER/Desktop/OPG-segmentation/Part 2/final/Dental_Dissertation/Report Generation/Dentist_Evaluation.xlsx'
EXCEL_HEADERS = [
    "File name", "Image Quality", "Report Quality", "Correctness", 
    "Completeness", "Relevance", "Clarity", "Feedback"
]

GROQ_API_KEYS, i = [], 1
while True:
    key = os.getenv(f"GROQ_API_KEY_{i}")
    if key:
        GROQ_API_KEYS.append(key)
        i += 1
    else:
        break

def get_groq_response(client, system_prompt: str, user_prompt: str) -> str:
    """Sends the prompt to Groq and handles errors."""
    try:
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
        chat_completion = client.chat.completions.create(
            messages=messages, model=MODEL_ID, temperature=0.7, max_tokens=2048
        )
        return chat_completion.choices[0].message.content
    except GroqError as e:
        if e.status_code == 429:
            if "tokens per day (tpd)" in str(e.message).lower():
                return "ERROR: Rate limit hit (TPD)"
            else:
                print("  [Info] Per-minute limit hit. Sleeping for 10 seconds...")
                time.sleep(10)
                return "ERROR: Retrying (RPM)"
        else:
            return f"ERROR: API call failed. {e}"
    except Exception as e:
        return f"ERROR: An unexpected error occurred. {e}"


# LLM <think> Tag Cleaning Function
def clean_llm_output(raw_text: str) -> str:
    """
    Strips the <think>...</think> block from the start of an LLM response.
    """
    match = re.search(r'</think>(.*)', raw_text, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return raw_text.strip()


# Markdown to ReportLab Conversion Function
def convert_markdown_to_flowables(md_text: str, styles: dict) -> list:
    """
    Converts a Markdown text string into a list of ReportLab Flowables.
    """
    flowables = []
    
    # Split the text by lines
    lines = md_text.split('\n')
    
    for line in lines:
        line = line.strip()
        
        # Handle Headings
        if line.startswith("### "):
            text = line.replace("### ", "").strip()
            flowables.append(Paragraph(text, styles['Heading3']))
        elif line.startswith("## "):
            text = line.replace("## ", "").strip()
            flowables.append(Paragraph(text, styles['Heading2']))
        elif line.startswith("# "):
            text = line.replace("# ", "").strip()
            flowables.append(Paragraph(text, styles['Heading1']))
        
        # Handle Horizontal Lines
        elif line in ["---", "***", "___"]:
            flowables.append(HRFlowable(width="100%", thickness=0.5, color=black, spaceAfter=12))
        
        # Handle Bullet Points
        elif line.startswith("* "):
            text = "&bull; " + line[2:].strip()
            # Apply inline bold
            text = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
            flowables.append(Paragraph(text, styles['Bullet']))
        
        # Handle Numbered Lists (e.g., "1. ...")
        elif re.match(r'^\d+\.\s+', line):
            # Apply inline bold
            line = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', line)
            flowables.append(Paragraph(line, styles['BodyText']))

        # Handle Empty Lines (as spacers)
        elif not line:
            flowables.append(Spacer(1, 0.25 * cm))
            
        # Handle normal text
        else:
            # Apply inline bold
            line = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', line)
            flowables.append(Paragraph(line, styles['BodyText']))
            
    return flowables


# PDF GENERATION FUNCTION
def create_pdf_report(image_path: str, caption_text: str, report_text: str, pdf_save_path: str):
    """
    Creates a PDF report with a single-column (stacked) layout.
    Converts markdown in the report_text to formatted PDF.
    """
    try:
        # Create document template
        page_width, page_height = portrait(A4)
        doc = SimpleDocTemplate(pdf_save_path, pagesize=portrait(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        
        # Add the Image
        if not os.path.exists(image_path):
            print(f"  [Warning] Image not found at {image_path}. Skipping PDF image.")
        else:
            with PILImage.open(image_path) as pil_img:
                img_width, img_height = pil_img.size
            max_img_width = page_width - (3 * cm) # Page width minus margins
            aspect = img_height / float(img_width)
            img_display_width = max_img_width
            img_display_height = max_img_width * aspect
            img = Image(image_path, width=img_display_width, height=img_display_height)
            story.append(img)
            story.append(Spacer(1, 0.5 * cm))

        # Create all necessary styles
        # styles = getSampleStyleSheet()
        # styles.add(ParagraphStyle(
        #     name='BodyText', parent=styles['Normal'], fontSize=9, leading=11
        # ))
        # styles.add(ParagraphStyle(
        #     name='CaptionText', parent=styles['Normal'], fontSize=8, leading=10
        # ))
        # styles.add(ParagraphStyle(
        #     name='Heading1', parent=styles['Heading1'], fontSize=14, leading=16, spaceAfter=6
        # ))
        # styles.add(ParagraphStyle(
        #     name='Heading2', parent=styles['Heading2'], fontSize=12, leading=14, spaceAfter=6
        # ))
        # styles.add(ParagraphStyle(
        #     name='Heading3', parent=styles['Heading3'], fontSize=10, leading=12, spaceAfter=4, fontName='Helvetica-Bold'
        # ))
        # styles.add(ParagraphStyle(
        #     name='Bullet', parent=styles['BodyText'], leftIndent=1*cm, firstLineIndent=-0.5*cm
        # ))
        styles = getSampleStyleSheet()

        # Modify BodyText
        styles['BodyText'].fontSize = 9
        styles['BodyText'].leading = 11
        
        # Add a NEW style for CaptionText
        styles.add(ParagraphStyle(
            name='CaptionText', 
            parent=styles['Normal'], 
            fontSize=8, 
            leading=10
        ))
        
        # Modify Heading1
        styles['Heading1'].fontSize = 14
        styles['Heading1'].leading = 16
        styles['Heading1'].spaceAfter = 6
        
        # Modify Heading2
        styles['Heading2'].fontSize = 12
        styles['Heading2'].leading = 14
        styles['Heading2'].spaceAfter = 6
        
        # Modify Heading3
        styles['Heading3'].fontSize = 10
        styles['Heading3'].leading = 12
        styles['Heading3'].spaceAfter = 4
        styles['Heading3'].fontName = 'Helvetica-Bold'
        
        # Modify Bullet
        styles['Bullet'].parent = styles['BodyText']
        styles['Bullet'].leftIndent = 1*cm
        styles['Bullet'].firstLineIndent = -0.5*cm

        

        # Add Grounding Caption (Title + Text)
        story.append(Paragraph("Grounding Caption", styles['Heading2']))
        story.append(Paragraph(caption_text.replace('\n', '<br/>'), styles['CaptionText']))
        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", thickness=1, color=black, spaceAfter=1, spaceBefore=1))
        story.append(Spacer(1, 0.5 * cm))

        # Add Medical Report (Title + Flowables)
        story.append(Paragraph("Medical Report", styles['Heading1']))
        report_flowables = convert_markdown_to_flowables(report_text, styles)
        story.extend(report_flowables)

        doc.build(story)
        return True
    except Exception as e:
        print(f"\n  [Error] Failed to create PDF for {os.path.basename(pdf_save_path)}: {e}")
        return False


def setup_evaluation_excel():
    """Loads or creates the evaluation Excel file."""
    existing_filenames = set()
    if os.path.exists(EVALUATION_EXCEL_FILE):
        print(f"Loading existing evaluation file: {EVALUATION_EXCEL_FILE}")
        wb = load_workbook(EVALUATION_EXCEL_FILE)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                existing_filenames.add(cell_value)
    else:
        print(f"Creating new evaluation file: {EVALUATION_EXCEL_FILE}")
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
    return wb, ws, existing_filenames


def check_all_images_exist(image_data_list, image_dir):
    """Checks if all images in the JSON exist in the source directory."""
    print("Start Verification Check")
    print(f"Checking for {len(image_data_list)} images in: {image_dir}")
    missing_files = []
    
    for img_data in tqdm(image_data_list, desc="Verifying images"):
        image_name = img_data.get("image_name")
        if not image_name:
            continue
            
        full_path = os.path.join(image_dir, image_name)
        
        if not os.path.exists(full_path):
            if image_name.endswith('.jpg.jpg'):
                image_name = image_name.replace('.jpg.jpg', '.jpg')
                img_data['image_name'] = image_name 
                full_path = os.path.join(image_dir, image_name)

            if not os.path.exists(full_path):
                 missing_files.append(image_name)

    if not missing_files:
        print(f"Success! All {len(image_data_list)} images are present.")
        return True, []
    else:
        print(f"\n--- !!! ERROR: VERIFICATION CHECK FAILED !!! ---")
        print(f"Found {len(missing_files)} missing image files. Cannot proceed.")
        for f in missing_files:
            print(f"  - Missing: {f}")
        return False, missing_files


def main():
    print(f"--- Medical Report Generation (Key Rotation & Excel Log) ---")
    print(f"Using model: {MODEL_ID}")

    if not GROQ_API_KEYS:
        print("Error: No API keys found. Please set environment variables.")
        return
        
    api_key_queue = deque(GROQ_API_KEYS)
    print(f"Loaded {len(api_key_queue)} API keys from environment.")

    try:
        with open(INPUT_JSON_FILE, 'r', encoding='utf-8') as f:
            all_image_data = json.load(f)
    except FileNotFoundError:
        print(f"Error: Input file not found at {INPUT_JSON_FILE}")
        return

    all_files_ok, _ = check_all_images_exist(all_image_data, SOURCE_IMAGE_DIR)
    if not all_files_ok:
        print("Stopping due to missing files.")
        return

    wb, ws, existing_filenames_in_excel = setup_evaluation_excel()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(PDF_OUTPUT_DIR, exist_ok=True)
    print(f"TXT Reports will be saved to: {OUTPUT_DIR}")
    print(f"PDF Reports will be saved to: {PDF_OUTPUT_DIR}")

    images_to_process = []
    total_synced_to_excel = 0
    
    print("Checking for existing PDFs to sync with Excel log...")
    for index, img_data in enumerate(all_image_data, start=1):
        image_name = img_data.get("image_name", "unknown_image")
        base_name = os.path.splitext(image_name)[0]
        pdf_name = f"{index}_{base_name}_report.pdf"
        pdf_save_path = os.path.join(PDF_OUTPUT_DIR, pdf_name)
        
        if os.path.exists(pdf_save_path):
            if pdf_name not in existing_filenames_in_excel:
                ws.append([pdf_name])
                existing_filenames_in_excel.add(pdf_name)
                total_synced_to_excel += 1
        else:
            images_to_process.append((index, img_data, pdf_name, pdf_save_path))
            
    if total_synced_to_excel > 0:
        print(f"Synced {total_synced_to_excel} existing PDFs to the Excel log.")
    
    wb.save(EVALUATION_EXCEL_FILE)
    
    total_to_process = len(images_to_process)
    total_already_exist = len(all_image_data) - total_to_process
    
    print(f"Found {len(all_image_data)} total images.")
    print(f"{total_already_exist} reports already exist.")
    print(f"{total_to_process} new reports remaining to generate.")
    
    if total_to_process == 0:
        print("All reports are already generated and logged. Exiting.")
        wb.close()
        return

    current_client = None
    current_key = None
    
    def get_new_client():
        if not api_key_queue:
            return None, None
        key = api_key_queue.popleft()
        print(f"\nSwitching to new API key: ...{key[-4:]} ({len(api_key_queue)} keys remaining)")
        return Groq(api_key=key), key

    current_client, current_key = get_new_client()
    if current_client is None:
        print("No valid API keys loaded from environment.")
        wb.close()
        return

    system_prompt = config_and_prompts.SYSTEM_PROMPT
    captions_created = 0
    reports_generated = 0
    pdfs_created = 0
    errors_encountered = 0
    
    image_index = 0
    with tqdm(total=total_to_process, desc="Generating Reports") as pbar:
        while image_index < total_to_process:
            
            index, image_data, pdf_name, pdf_save_path = images_to_process[image_index]
            image_name = image_data.get("image_name", "unknown_image")
            base_name = os.path.splitext(image_name)[0]
            
            caption_path = os.path.join(OUTPUT_DIR, f"{base_name}_input_caption.txt")
            report_txt_path = os.path.join(OUTPUT_DIR, f"{base_name}_medical_report.txt")
            source_image_path = os.path.join(SOURCE_IMAGE_DIR, image_name)

            grounding_caption = generate_grounding_caption.create_grounding_caption(image_data)
            with open(caption_path, 'w', encoding='utf-8') as f:
                f.write(grounding_caption)
            captions_created += 1

            medical_report_raw = get_groq_response(current_client, system_prompt, grounding_caption)
            
            # Clean the LLM output FIRST
            medical_report = clean_llm_output(medical_report_raw)

            if medical_report_raw.startswith("ERROR:"): # Check the *raw* response for errors
                if medical_report_raw == "ERROR: Rate limit hit (TPD)":
                    pbar.set_description(f"Key ...{current_key[-4:]} exhausted. Rotating...")
                    current_client, current_key = get_new_client()
                    if current_client is None:
                        print("\nAll API keys are exhausted for the day. Stopping script.")
                        break
                
                elif medical_report_raw == "ERROR: Retrying (RPM)":
                    pbar.set_description(f"Per-minute limit. Retrying image {image_name}...")
                    
                else:
                    print(f"\nNon-retryable error for {image_name}: {medical_report_raw}")
                    with open(report_txt_path, 'w', encoding='utf-8') as f:
                        f.write(medical_report_raw)
                    reports_generated += 1
                    errors_encountered += 1
                    image_index += 1
                    pbar.update(1)
            else:
                # SUCCESS
                with open(report_txt_path, 'w', encoding='utf-8') as f:
                    f.write(medical_report) # Save the *clean* report
                reports_generated += 1
                
                pdf_success = create_pdf_report(
                    image_path=source_image_path,
                    caption_text=grounding_caption,
                    report_text=medical_report, # Pass the *clean* report
                    pdf_save_path=pdf_save_path
                )
                
                if pdf_success:
                    pdfs_created += 1
                    ws.append([pdf_name]) # Add to Excel
                
                image_index += 1
                pbar.update(1)
                
                if image_index % 10 == 0:
                    wb.save(EVALUATION_EXCEL_FILE)

    # Final Summary Log
    print(f"\n--- Pipeline Complete (or all keys used) ---")
    
    try:
        wb.save(EVALUATION_EXCEL_FILE)
        print(f"Evaluation log saved: {EVALUATION_EXCEL_FILE}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
    wb.close()

    print(f"\n--- FINAL SUMMARY ---")
    print(f"Total images in JSON:           {len(all_image_data)}")
    print(f"Reports skipped (already exist): {total_already_exist}")
    print(f"Reports synced to Excel log:     {total_synced_to_excel}")
    print("---")
    print(f"--- Processing in This Run ---")
    print(f"Grounding captions created:      {captions_created}")
    print(f"LLM text reports created:        {reports_generated}")
    print(f"PDF reports created:             {pdfs_created}")
    print(f"Errors encountered:              {errors_encountered}")
    print("---")
    print(f"Total reports remaining:         {total_to_process - image_index}")

if __name__ == "__main__":
    main()